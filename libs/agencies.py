# -*- coding: utf-8 -*-
import json
import requests
import ssl
from bs4 import BeautifulSoup
import calendar
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.poolmanager import PoolManager
from requests.packages.urllib3.util.ssl_ import create_urllib3_context
from openpyxl import load_workbook
import xlrd


debug = False


class TLSv1Adapter(HTTPAdapter):
    """"Transport adapter" that allows us to use TLSv1."""

    def init_poolmanager(self, connections, maxsize, block=False):
        self.poolmanager = PoolManager(
                num_pools=connections,
                maxsize=maxsize,
                block=block,
                ssl_version=ssl.PROTOCOL_TLSv1)


class NODHAdapter(HTTPAdapter):
    def init_poolmanager(self, *args, **kwargs):
        context = create_urllib3_context(
            ciphers='ECDH+AESGCM:DH+AESGCM:ECDH+AES256:DH+AES256:ECDH+AES128:DH+AES:ECDH+HIGH:!DH')
        kwargs['ssl_context'] = context
        return super(NODHAdapter, self).init_poolmanager(*args, **kwargs)


# Gather data from various agencies depending on country code of tail number

class Owner:
    def __init__(self):
        self.name   = 'TBD'
        self.country= 'TBD'

    def __init__(self, name, street, city, zip_code, country):
        self.name       = name
        self.street     = street
        self.city       = city
        self.zip_code   = zip_code
        self.country    = country

    def __repr__(self):
        return "Name: %s \n Street: %s\n City: %s\n ZIP: %s\n Country: %s" % \
            (self.name, self.street, self.city, self.zip_code, self.country)

    def __str__(self):
        return self.__repr__()


def CH(tail_n):
    """
        Get information on aircraft from tail number
    """
    url = 'https://app02.bazl.admin.ch/web/bazl-backend/lfr'
    headers = {
        'Pragma': 'no-cache',
        'Origin': 'https://app02.bazl.admin.ch/web/bazl/fr/',
        'Content-Type': 'application/json',
        'Referer': 'https://app02.bazl.admin.ch/web/bazl/fr/'
    }
    data = '{"page_result_limit":10,'\
        '"current_page_number":1,'\
        '"sort_list":"registration",'\
        '"language":"fr",'\
        '"queryProperties":{"registration":"'+tail_n[3:] + '",'\
        '"aircraftStatus":'\
        '["Registered","Reserved","Reservation Expired","Registration in Progress"]'\
        '}}}'

    response = requests.post(url, headers=headers, data=data)

    if response.status_code == 200:
        jsonobj = json.loads(response.text)
        if len(jsonobj) == 0:
            print("[!][CH][{}] Error when retrieving from registry".format(tail_n))
            return
        infoarray = jsonobj[0]
        own_ops = infoarray.get('ownerOperators')
        leng = len(own_ops)
        name = own_ops[leng-1].get('ownerOperator')
        addr = own_ops[leng-1].get('address')
        street = addr.get('street')
        street_n = addr.get('streetNo')
        zipcode = addr.get('zipCode')
        city = addr.get('city')
        owner = Owner(name, street + ' ' + street_n,
                    city, zipcode, "Switzerland")
        return owner
    else:
        print("[!] Error retrieving from CH")


def FR(tail_n):
    s = requests.session()
    headers = {
        'Origin': 'https://immat.aviation-civile.gouv.fr',
        'Referer': 'https://immat.aviation-civile.gouv.fr/immat/servlet/aeronef_liste.html'
    }

    data = [
        ('FORM_DTO_ID', 'DTO_RECHERCHE_AER'),
        ('FORM_ACTION', 'SEARCH'),
        ('FORM_CHECK', 'true'),
        ('FORM_ROW', ''),
        ('CTRL_ID', '1010'),
        ('SUBFORM_DTC_ID', 'DTC_1'),
        ('$DTO_RECHERCHE_AER$PER_ID', ''),
        ('$DTO_RECHERCHE_AER$PER_VERSION', ''),
        ('$DTO_RECHERCHE_AER$PERSONNE', ''),
        ('$DTO_RECHERCHE_AER$MRQ_CD', tail_n),
        ('$DTO_RECHERCHE_AER$CNT_LIBELLE', ''),
        ('$DTO_RECHERCHE_AER$SI_RADIE', '1'),
        ('$DTO_RECHERCHE_AER$SI_RADIEcheckbox', '1'),
        ('$DTO_RECHERCHE_AER$SI_INSCRIT', '1'),
        ('$DTO_RECHERCHE_AER$SI_INSCRITcheckbox', '1'),
        ('$DTO_RECHERCHE_AER$CRE_NUM_SERIE', ''),
    ]
    #s.mount("https://", TLSv1Adapter()) # Add TSLV1 adapter (outdated)

    response = s.post(
        'https://immat.aviation-civile.gouv.fr/immat/servlet/aeronef_liste.html',
        headers=headers,
        data=data)

    if response.status_code == 200:
        soup    = BeautifulSoup(response.text, 'html.parser')
        bls     = soup.find('a', string=tail_n)

        if bls is None:
            print('[!][FR] Could not find supplied tail number in agency registers')
        else:
            response = s.get(
                'https://immat.aviation-civile.gouv.fr/immat/servlet/' + bls['href'], 
                headers=headers)
            soup    = BeautifulSoup(response.text, 'html.parser')
            bls     = soup.find('a', string="Données juridiques")
            r       = s.get('https://immat.aviation-civile.gouv.fr/immat/servlet/' + bls['href'])
            if r.status_code == 200:
                soup = BeautifulSoup(r.text, 'html.parser')
                bls  = soup.find_all('td', {'class': "tdLigneListe"})
                if len(bls) > 0:
                    name = bls[0].text
                    addr = bls[1].text
                    city = bls[2].text
                    return Owner(name, addr, city, '', 'France')
                else:
                    print('[!][FR] Error while retrieving info for {}'.format(tail_n))
            else:
                print('[!][FR] Error while retrieving info for {}'.format(tail_n))

def US(tail_n):
    name = ''
    addr = ''
    city = ''

    resp = requests.get(
        "https://registry.faa.gov/AircraftInquiry/Search/NNumberResult?nNumberTxt="+tail_n)
    if resp.status_code == 200:
        soup = BeautifulSoup(resp.content, 'html.parser')

        # Get table regarding owner information
        tables = soup.find_all(
                'table', {'class', 'devkit-table'})
        for table in tables:
            caption = table.find('caption', {'class', 'devkit-table-title'})

            # This is the table we are interested in
            if caption.text == 'Registered Owner':
                print(table)
                rows = table.find_all('tr')
                for row in rows:
                    cols = row.find_all('td')
                    for col in cols:
                        if col['data-label'] == 'Name':
                            name = col.text
                        elif col['data-label'] == 'Street':
                            addr = addr + col.text
                        elif col['data-label'] == 'City':
                            city = col.text
                        elif col['data-label'] == 'State':
                            city = city + ', ' + col.text
                        elif col['data-label'] == 'Zip Code':
                            zip_code = col.text
        return Owner(name, addr, city, zip_code, 'USA')
    else:
        print("[!][{}] HTTP status code from {}"\
                .format(resp.status_code, resp.url))
    print('[!] Error while retrieving from US')


def IS(tail_n):
    name = ''
    street = ''
    city = ''
    s = requests.session()
    # s.mount("https://", TLSv1Adapter()) # Outdated TSLv1 -> Needs a specific adapter
    req = s.get(
        'https://www.icetra.is/aviation/aircraft/register?aq='+tail_n)
    if req.status_code == 200:
        soup = BeautifulSoup(req.text, 'html.parser')
        own = soup.find('li', {'class': 'owner'})
        if own is not None:
            won = own.stripped_strings
            for i, j in enumerate(won):
                if i == 1:
                    name = j
                elif i == 2:
                    street = j
                elif i == 3:
                    city = j
        if own is not None:
            return Owner(name, street, city, '', 'Iceland')


def BE(tail_n):
    rep = requests.get('https://es.mobilit.fgov.be/aircraft-registry/rest/aircrafts?aircraftStates=REGISTERED&registrationMark=' +
                       tail_n + '&page=0&pageSize=10&sort=REGISTRATIONMARK&sortDirection=ascending')

    if rep.status_code == 200:
        j = json.loads(rep.text)
        if len(j) > 0:
            if j[0].get('id') != '':
                rep = requests.get(
                    'https://es.mobilit.fgov.be/aircraft-registry/rest/aircrafts/'+str(j[0].get('id')))
                if rep.status_code == 200:
                    # print(rep.text)
                    j = json.loads(rep.text)
                    name = j.get('stakeHolderRoleList')[0].get('name')
                    street = j.get('stakeHolderRoleList')[
                        0].get('addresses').get('street')
                    city = j.get('stakeHolderRoleList')[
                        0].get('addresses').get('city')
                    return Owner(name, street, city, '', 'Belgium')
                else:
                    return Exception("Error retrieving from BE")


def SW(tail_n):
    data = {
        'selection': 'regno',
        'regno': tail_n,
        'part': '',
        'owner': '',
        'item': '',
        'X-Requested-With': 'XMLHttpRequest'
    }
    s = requests.session()
    rep = s.post(
        'https://sle-p.transportstyrelsen.se/extweb/en-gb/sokluftfartyg', data=data)
    if rep.status_code == 200:
        soup = BeautifulSoup(
            rep.content, features="html.parser")
        results_element = soup.find_all(
            'label', {'class': 'owner-headline'})[0]
        name = results_element.parent.find(
            'a', {'class': 'open-owner-page'}).text
        street = results_element.parent.text.split('\r\n')[
            1].strip()
        city = results_element.parent.text.split('\r\n')[
            2].strip()
        country = results_element.parent.text.split('\r\n')[
            3].strip()
        return Owner(name, street, city, '', country)
    else:
        raise Exception('Error retrieving from SW')


def AT(tail_n):
    tail_n = tail_n.lstrip('OE-')
    rep = requests.get(
            'https://www.austrocontrol.at/'\
                    'lfa-publish-service/v2/oenfl/'\
                    'luftfahrzeuge?kennzeichen='+tail_n
            )
    print(rep.url)

    if rep.status_code == 200:
        j = json.loads(rep.content)
        owner_infos = [x for x in j if x['kennzeichen'] == tail_n]
        name, loc_info, country = owner_infos[0]['halter'].split('\r\n')
        city, street  = loc_info.split(', ')
        npa, city = city.split(' ')
        return Owner(name, street, city, npa, 'Austria')

def CZ(tail_n):
    data = {
        'aparameters': [
            '',
            '',
            '',
            '',
            '',
            tail_n[
                3:],
            '',
            '',
            '',
            '',
            '',
            'kategorie_letadla:',
            'typ_letadla:',
            'rejstrikova_znacka:' +
            str(
                tail_n[3:]),
            'aid_nast:281',
            'aSubmit:1',
        ]
    }

    r = requests.post(
        'http://portal.caa.cz/web_redir', data=data)
    soup = BeautifulSoup(
        r.content, features="html.parser")
    res = soup.find(
        'div', {'id': 'infobox_letadlo1'})
    name = res.find(
        'div', {'class': 'value'}).text
    return Owner(name, '', '', '', '')

def UK(tail_n):
    raise NotImplementedError(
        'UK registry not yet implemented. Registry url is https://siteapps.caa.co.uk/g-info/')
    data = {
        'Registration': tail_n[2:]
    }
    print(
        data)
    s = requests.session()
    r = s.get(
        'https://siteapps.caa.co.uk/g-info/')
    r = s.post(
        'https://ginfoapi.caa.co.uk/api/aircraft/search', data=data)
    if r.status_code == 200:
        print(r.content)

def IE(tail_n):
    headers = {
        'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64; rv:68.0) Gecko/20100101 Firefox/68.0'}
    # r 		= requests.get('https://www.iaa.ie/commercial-aviation/aircraft-registration-2/latest-register-and-monthly-changes-1', headers=headers)
    url_doc = 'https://www.iaa.ie/docs/default-source/publications/aircraft-registration/30-november-2019.xlsx'
    r = requests.get(
        url_doc, headers=headers)
    if r.status_code == 200:
        with open('/tmp/book.xlsx', 'wb') as f:
            f.write(
                r.content)
            book = load_workbook(
                '/tmp/book.xlsx')
            for sheet in book:
                plane = [
                    row for row in sheet.values if row[0] == tail_n]
                if len(plane) > 0:
                    name = plane[
                        0][12]
                    addr = plane[
                        0][13]
                    return Owner(name, addr, '', '', 'Ireland')
                    raise Exception(
                        'Error retrieving from Ireland register. Tail number may be wrong')


def IM(tail_n):
    tail_n = tail_n.lstrip('M-')
    r = requests.get('https://ardis.iomaircraftregistry.com/register/search?prs_rm__ptt=1&prs_rm__tt=1&prs_as__v=2&prs_rm__v1='+tail_n+'&prs_rm__pv1='+tail_n)
    if r.status_code == 200:
        soup = BeautifulSoup(
            r.content, features="html.parser")
        link = soup.find(
            'td', {'id': 'prp__rid__1__cid__2'})
        if link is not None:
            href = link.contents[0]['href']
            r = requests.get(
                'https://ardis.iomaircraftregistry.com'+href)
            if r.status_code == 200:
                soup = BeautifulSoup(
                    r.content, features="html.parser")
                own = soup.find(
                    'span', {'id': 'prv__11__value'})
                name, infos = own.text.split(',')
                if len(infos) > 1:
                    street = infos
                    return Owner(name, street, '', '', '')

def CA(tail_n):
    data = {
        '__EVENTTARGET': '',
        '__EVENTARGUMENT': '',
        'ctl00$ContentPlaceHolder1$btnSearchTop': 'Search',
        'ctl00$ContentPlaceHolder1$txtMark': tail_n[2:],
        'ctl00$ContentPlaceHolder1$ddlMark': '0',
        'ctl00$ContentPlaceHolder1$txtCommonName': '',
        'ctl00$ContentPlaceHolder1$ddlCommonName': '0',
        'ctl00$ContentPlaceHolder1$txtModelName': '',
        'ctl00$ContentPlaceHolder1$ddlModelName': '0',
        'ctl00$ContentPlaceHolder1$txtSerialNumber': '',
        'ctl00$ContentPlaceHolder1$ddlSerialNumber': '0',
        'ctl00$ContentPlaceHolder1$txtWeightFrom': '',
        'ctl00$ContentPlaceHolder1$txtWeightTo': '',
        'ctl00$ContentPlaceHolder1$ddlImportYear': '%%',
        'ctl00$ContentPlaceHolder1$ddlAircraftCategory': '%%',
        'ctl00$ContentPlaceHolder1$ddlNumberOfEngines': '%%',
        'ctl00$ContentPlaceHolder1$ddlEngineCategory': '%%',
        'ctl00$ContentPlaceHolder1$txtIndustry': '',
        'ctl00$ContentPlaceHolder1$ddlIndustry': '0',
        'ctl00$ContentPlaceHolder1$ddlAssemblyYear': '%%',
        'ctl00$ContentPlaceHolder1$ddlAssemblyCountry': '%%',
        'ctl00$ContentPlaceHolder1$txtOwnerName': '',
        'ctl00$ContentPlaceHolder1$ddlOwnerName': '0',
        'ctl00$ContentPlaceHolder1$txtTradeName': '',
        'ctl00$ContentPlaceHolder1$ddlTradeName': '0',
        'ctl00$ContentPlaceHolder1$txtCity': '',
        'ctl00$ContentPlaceHolder1$ddlCity': '0',
        'ctl00$ContentPlaceHolder1$ddlRegion': '%%',
        'ctl00$ContentPlaceHolder1$ddlProvince': '%%',
        'ctl00$ContentPlaceHolder1$txtPostalCode': '',
        'ctl00$ContentPlaceHolder1$ddlPostalCode': '0',
        'ctl00$ContentPlaceHolder1$ddlMultipleOwner': 'A',
        'ctl00$ContentPlaceHolder1$ddlRegionalOffice': '%%',
        'ctl00$ContentPlaceHolder1$ddlRegistrationType': '%%',
        'ctl00$ContentPlaceHolder1$ddlRegistrationEligibility': '%%',
    }

    s = requests.session()
    s.mount(
        'https://', NODHAdapter())
    r = s.get(
        'https://wwwapps.tc.gc.ca/Saf-Sec-Sur/2/CCARCS-RIACC/RchAvc.aspx')
    if r.status_code == 200:
        soup = BeautifulSoup(
            r.text, features='html.parser')
        input_viewstate_gen = soup.find(
            'input', {'id': '__VIEWSTATEGENERATOR'})
        input_viewstate = soup.find(
            'input', {'id': '__VIEWSTATE'})
        input_validation = soup.find(
            'input', {'id': '__EVENTVALIDATION'})
        data['__EVENTVALIDATION'] = input_validation[
            'value']
        data['__VIEWSTATEGENERATOR'] = input_viewstate_gen[
            'value']
        data['__VIEWSTATE'] = input_viewstate[
            'value']
        headers = {
            'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64; rv:68.0) Gecko/20100101 Firefox/68.0'}
        r = s.post(
            'https://wwwapps.tc.gc.ca/Saf-Sec-Sur/2/CCARCS-RIACC/RchAvc.aspx', data=data, headers=headers)
        soup = BeautifulSoup(
            r.text, features="html.parser")
        name = soup.find('div', {'id': 'dvOwnerName'}).text.replace(
            'Name:', '').strip()
        city = soup.find('div', {'id': 'divOwnerCity'}).find(
            'div', {'class': 'span-4'}).text.replace('City:', '').strip()
        addr = soup.find('div', {'id': 'divOwnerAddress'}).text.replace(
            'Address:', '').strip()
        return Owner(name, addr, city, '', 'Canada')

def DE(tail_n):
    raise NotImplementedError(
        'German register not implemented yet')

def IT(tail_n):
    headers = {
        'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64; rv:68.0) Gecko/20100101 Firefox/68.0'}
    s = requests.Session()
    r = s.get(
        'http://gavs.it/')
    soup = BeautifulSoup(
        r.content, features="html.parser")
    csrf = soup.find(
        'input', {'name': 'csrf_test_name'})
    data = {
        'csrf_test_name': csrf['value'],
        'registration': tail_n[2:]
    }

    r = s.post(
        'https://gavs.it/rci/search_registration', 
        data=data, 
        headers=headers)
    if r.status_code == 200:
        record_url = r.headers['Refresh'].split(';')[1][4:]
        r = s.get(
            record_url)
        if r.status_code == 200:
            soup = BeautifulSoup(
                r.text, features="html.parser")
            tab_owners = soup.find(
                'div', {'id': 'htab2'})
            tab_owner = tab_owners.find_all(
                'dl', {'class': 'dl-horizontal'})[0]
            owner_name = tab_owner.find_all(
                    'dd')[-1].text
            return Owner(owner_name, '', '', '', '')
        raise Exception(
                "Could not get info from IT register")

def RO(tail_n):
    raise NotImplementedError(
        'Romanian register is the following pdf document http://www.caa.ro/media/docs/OPERATORI_AERIENI_ROMANI_28.11.2019_rom-eng.pdf. Documents parsing not implemented yet')

def HR(tail_n):
    raise NotImplementedError(
        'Croatian register is a pdf document. The document is available at https://www.ccaa.hr/english/popis-registriranih-zrakoplova_101/')

def AU(tail_n):
    """
    Australia Registry

    """
    r = requests.get(
        'https://www.casa.gov.au/search-centre/aircraft-register/'+tail_n[3:])
    if r.status_code == 200:
        soup = BeautifulSoup(
            r.content, features="html.parser")
        
        name = soup.find('div', {'class':'field--name-field-registration-holder'}).text.strip('Registration holder:\n')
        addr = soup.find('div', {'class':'field--name-field-reg-hold-address-1'}).text.strip('Address 1:\n')
        city = soup.find('div', {'class':'field--name-field-tx-reg-hold-suburb'}).text.strip('Suburb / City:\n')
        return Owner(name, addr, city, '', 'Australia')
    raise Exception(
            "Could not get info from AU register")

def SG(tail_n):
    raise NotImplementedError('Singapore register is a pdf document. The document is available at https://www.caas.gov.sg/docs/default-source/pdf/singapore-registered-aircraft-engine-nos---oct-2019.pdf')

def NZ(tail_n):
    r = requests.get('https://www.aviation.govt.nz/aircraft/aircraft-registration/aircraft-register/ShowDetails/'+tail_n[3:])
    if r.status_code == 200:
        soup		= BeautifulSoup(r.text, features="html.parser")
        owner_info 	= soup.find('div', {'class': 'col-md-9'})
        name 		= owner_info.text.strip().split('\n')[1].strip()
        street 		= owner_info.text.strip().split('\n')[2].strip()
        return Owner(name, street, '', '', 'New Zealand')
    raise Exception("Could not get info from NZ register")

def BR(tail_n):
    r = requests.get('https://sistemas.anac.gov.br/aeronaves/cons_rab_resposta.asp?textMarca=' + tail_n)
    if r.status_code == 200:
        soup 		= BeautifulSoup(r.text, features="html.parser")
        headings 	= soup.find_all('th', {'scope':'row'})
        for heading in headings:
            if 'Proprietário' in heading.text:
                name = heading.parent.td.text.strip()
                return Owner(name, '', '', '', 'Brazil')
            raise Exception("Could not get info from BR register. " + r.url + r.status_code )

def UA(tail_n):
    r = requests.get('http://avia.gov.ua/register_of_aircraft.xls')
    if r.status_code == 200:
        with open('/tmp/register.xls', 'wb') as f:
            f.write(r.content)
            book 		= xlrd.open_workbook('/tmp/register.xls')
            sheet_names	= book.sheet_names()
            xl_sheet 	= book.sheet_by_name(sheet_names[0])
            for i in range(0, xl_sheet.nrows):
                if xl_sheet.row(i)[2].value == tail_n:
                    name = xl_sheet.row(i)[9].value
                    return Owner(name, '', '', '', 'Ukraine')
                    raise Exception("Could not get info from UA register")

def TH(tail_n):
    # Register dates from 2019. TODO: find more recent one
    r = requests.get('https://www.caat.or.th/wp-content/uploads/2019/03/Aircraft-List_18-Mar-2019-Un-controlled-EN.xlsx')
    if r.status_code == 200:
        with open('/tmp/book.xlsx', 'wb') as f:
            f.write(r.content)
            book = load_workbook('/tmp/book.xlsx')
            for sheet in book:
                infos = [row for row in sheet.values if row[1] == tail_n]
                if len(infos) > 0:
                    name = infos[0][2]
                    return Owner(name, '', '', '', 'Thailand')
                else:
                    raise Exception('Number not found in thai register')
            raise Exception("Could not get info from thai register")
