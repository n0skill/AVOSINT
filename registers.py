# -*- coding: utf-8 -*-

import json
import requests
import ssl
import calendar
import urllib

from bs4 import BeautifulSoup
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.poolmanager import PoolManager
from requests.packages.urllib3.util.ssl_ import create_urllib3_context
from urllib3.exceptions import NewConnectionError
from openpyxl import load_workbook
from parsr_client import ParsrClient as client
from pprint import pprint
from aircraft import Aircraft

debug = False

from urllib3.exceptions import InsecureRequestWarning
# Suppress only the single warning from urllib3 needed.
requests.packages.urllib3.disable_warnings(category=InsecureRequestWarning)

class TLSv1Adapter(HTTPAdapter):
    """"Transport adapter" that allows us to use TLSv1."""

    def init_poolmanager(self, connections, maxsize, block=False):
        self.poolmanager = PoolManager(
                num_pools=connections,
                maxsize=maxsize,
                block=block,
                ssl_version=ssl.PROTOCOL_TLSv1)


class NODHAdapter(HTTPAdapter):
    def init_poolmanager(self, *args, **kwargs):
        context = create_urllib3_context(
            ciphers='ECDH+AESGCM:DH+AESGCM:ECDH+AES256:DH+AES256:ECDH+AES128:DH+AES:ECDH+HIGH:!DH')
        kwargs['ssl_context'] = context
        return super(NODHAdapter, self).init_poolmanager(*args, **kwargs)

#################################################
# Registers.py
# Goal: Gather data from various agencies 
# depending on country code of tail number
#
# Returns tuple: (owner infos, aircraft infos)
#


class Owner:
    def __init__(self):
        self.name   = 'TBD'
        self.country= 'TBD'

    def __init__(self, name, street='Unknown', city='Unknown', zip_code='Unknown', country='Unknown'):
        self.name       = name
        self.street     = street
        self.city       = city
        self.zip_code   = zip_code
        self.country    = country

    def __repr__(self):
        return u"""
            Owner Informations:
            ==================
            Name: %s
            Street: %s
            City: %s
            ZIP: %s
            Country: %s
            """ % (self.name, self.street, self.city, self.zip_code, self.country)

    def __str__(self):
        return self.__repr__()

class DataSource:
    def __init__(self, url, src_type, request_type, is_secure, http_data='', headers=''):
        self.url = url
        self.src_type = src_type
        self.request_type = request_type
        self.is_secure = is_secure
        self.data = http_data
        self.headers = headers

        if self.headers == '':
            self.headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_5) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/12.1.1 Safari/605.1.15'}

    def gather(self, tail_n):
        """
        Gathers data from the configured DataSource for a register
        Returns a parsable object:
        - workbook for xlsx sources
        - json object for json sources
        - json object for pdf sources (using parsr)
        - raw request response content when content type could not be determined
        """
        print('[*] Gathering info from url', self.url)
        if self.data != '' and self.data is not None:
            print('[*] Replacing {{TAILN}} from data with acutal tail number', self.url)
            self.data = self.data.replace('{{TAILN}}', tail_n)
        
        try:
            r = None
            # JSON
            if self.src_type == 'json':
                if self.request_type=="GET":
                    r = requests.get(self.url+tail_n,headers=self.headers)
                if self.request_type=="POST":
                    r = requests.post(self.url, data=self.data, headers=self.headers)
                if r.status_code == 200:
                    j = json.loads(r.content)
                return j

            # Online xls 
            elif self.src_type == 'xlsx':
                with open('/tmp/book.xlsx', 'wb') as f:
                    f.write(r.content)
                    book = load_workbook(
                        '/tmp/book.xlsx')
                return book

            # PDF file
            elif self.src_type == 'pdf':
                parsr = client('localhost:3001')
                r = requests.get(self.url)
                if r.status_code == 200:
                    with open('/tmp/avosint.pdf', 'wb') as f:
                        f.write(r.content)
                    job = parsr.send_document(
                        file_path='/tmp/avosint.pdf',
                        config_path='./parsr.conf',
                        document_name='Sample File2',
                        wait_till_finished=True,
                        save_request_id=True,
                        silent=False)
                    j = parsr.get_json()
                    return j

            # Google spreadsheet
            elif self.src_type == 'google_sheets':
                from google_auth_oauthlib.flow import InstalledAppFlow
                from googleapiclient.discovery import build
                from googleapiclient.errors import HttpError
                
                # If modifying these scopes, delete the file token.json.
                SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']
                # The ID and range of a sample spreadsheet.
                SAMPLE_SPREADSHEET_ID = '1Kgu0uoXLGhoCHUyMgDsdqtW_XH3fvIglkVx5EdJhRnU'
                SAMPLE_RANGE_NAME = 'Class Data!A2:E'
                 # Call the Sheets API
                service = build('sheets', 'v4', credentials=creds)
                sheet = service.spreadsheets()
                result = sheet.values().get(spreadsheetId=SAMPLE_SPREADSHEET_ID,
                                            range=SAMPLE_RANGE_NAME).execute()
                values = result.get('values', [])

                if not values:
                    print('No data found.')
                    return

                print('Name, Major:')
                for row in values:
                    # Print columns A and E, which correspond to indices 0 and 4.
                    print('%s, %s' % (row[0], row[4]))
                return 0

        except NewConnectionError as e:
            print("[!] Failed to establish new connection. Try again ?")
        except Exception as e:
            print("[!!!] Exception occurred", e)
            print("[*] Try again with session")
            s = requests.session()
            r = s.get(self.url, headers=self.headers)
            if self.src_type == 'pdf':
                parsr = client('localhost:3001')
                r = requests.get(self.url)
                if r.status_code == 200:
                    with open('/tmp/avosint.pdf', 'wb') as f:
                            f.write(r.content)
                    job = parsr.send_document(
                        file_path='/tmp/avosint.pdf',
                        config_path='./parsr.conf',
                        document_name='Sample File2',
                        wait_till_finished=True,
                        save_request_id=True,
                        silent=False)
                    j = parsr.get_json()
                    print(j)
                    return j
                
            print(r)



        else:
            print("[!] Error {} when retrieving from\n[!] URL: {}".format(r.status_code, self.url))
            print("[!] DataSource error while gathering information.")
            print("[!] Try to gather using session.")

class Registry:
    def __init__(self, 
            name, 
            url,
            data_source_url, 
            data_source_type,
            data_request_type,
            post_data='',
            headers='',
            is_secure=True,
            tail_start=0):

        self.name = name
        self.url = url
        self.data_source = DataSource(
                data_source_url,
                data_source_type,
                data_request_type,
                is_secure, 
                post_data,
                headers) 
        self.results = []
        self.tail_start = tail_start

    def request_infos(self, tail_n):
        # Format tail number according to regex
        tail_n = tail_n[self.tail_start:]
        res = self.data_source.gather(tail_n)
        return res


def register_from_config(key):
    with(open('./registers.json', 'r')) as f:
        j = json.load(f)

    config = j[key]
    generic_registry = Registry(
        config['name'],
        config['website'],
        config['data_url'],
        config['data_type'],
        config['request_type'],
        config['data'] if 'data' in config else None,
        headers=config['headers'] if 'headers' in config else None,
        tail_start=config['tail_index_for_search'])
    return generic_registry



def NL(tail_n):
    return

def CH(tail_n):
    """
        Get information on aircraft from tail number
    """
    SwissRegister = register_from_config("CH")
    jsonobj = SwissRegister.request_infos(tail_n)
    if len(jsonobj) == 0:
        print("[!][CH][{}] Error when retrieving from registry".format(tail_n))
        return
    
    infoarray = jsonobj[0]
    own_ops = infoarray.get('ownerOperators')
    leng = len(own_ops)
    name = own_ops[leng-1].get('ownerOperator')
    addr = own_ops[leng-1].get('address')
    street = addr.get('street')
    street_n = addr.get('streetNo')
    zipcode = addr.get('zipCode')
    city = addr.get('city')
    owner = Owner(name, street + ' ' + street_n,
                city, zipcode, "Switzerland")
    return owner, Aircraft(tail_n)


def FR(tail_n):
    s = requests.session()
    headers = {
        'Origin': 'https://immat.aviation-civile.gouv.fr',
        'Referer': 'https://immat.aviation-civile.gouv.fr/immat/servlet/aeronef_liste.html'
    }

    data = [
        ('FORM_DTO_ID', 'DTO_RECHERCHE_AER'),
        ('FORM_ACTION', 'SEARCH'),
        ('FORM_CHECK', 'true'),
        ('FORM_ROW', ''),
        ('CTRL_ID', '1010'),
        ('SUBFORM_DTC_ID', 'DTC_1'),
        ('$DTO_RECHERCHE_AER$PER_ID', ''),
        ('$DTO_RECHERCHE_AER$PER_VERSION', ''),
        ('$DTO_RECHERCHE_AER$PERSONNE', ''),
        ('$DTO_RECHERCHE_AER$MRQ_CD', tail_n),
        ('$DTO_RECHERCHE_AER$CNT_LIBELLE', ''),
        ('$DTO_RECHERCHE_AER$SI_RADIE', '1'),
        ('$DTO_RECHERCHE_AER$SI_RADIEcheckbox', '1'),
        ('$DTO_RECHERCHE_AER$SI_INSCRIT', '1'),
        ('$DTO_RECHERCHE_AER$SI_INSCRITcheckbox', '1'),
        ('$DTO_RECHERCHE_AER$CRE_NUM_SERIE', ''),
        ]
    #s.mount("https://", TLSv1Adapter()) # Add TSLV1 adapter (outdated)

    response = s.post('https://immat.aviation-civile.gouv.fr/immat/servlet/aeronef_liste.html',
        headers=headers,
        data=data)

    if response.status_code == 200:
        soup    = BeautifulSoup(response.text, 'html.parser')
        bls     = soup.find('a', string=tail_n)

        if bls is None:
            print('[!][FR] Could not find supplied tail number in agency registers')
        else:
            response = s.get(
                'https://immat.aviation-civile.gouv.fr/immat/servlet/' + bls['href'], 
                headers=headers)
            soup    = BeautifulSoup(response.text, 'html.parser')
            bls     = soup.find('a', string="Données juridiques")
            r       = s.get('https://immat.aviation-civile.gouv.fr/immat/servlet/' + bls['href'])
            if r.status_code == 200:
                soup = BeautifulSoup(r.text, 'html.parser')
                bls  = soup.find_all('td', {'class': "tdLigneListe"})
                if len(bls) > 0:
                    name = bls[0].text
                    addr = bls[1].text
                    city = bls[2].text
                    return Owner(name, addr, city, '', 'France'), Aircraft(tail_n)
                else:
                    print('[!][FR] Error while retrieving info for {}'.format(tail_n))
            else:
                print('[!][FR] Error while retrieving info for {}'.format(tail_n))

def US(tail_n):
    name = ''
    addr = ''
    city = ''

    resp = requests.get(
        "https://registry.faa.gov/AircraftInquiry/Search/NNumberResult?nNumberTxt="+tail_n)
    if resp.status_code == 200:
        soup = BeautifulSoup(resp.content, 'html.parser')

        # Get table regarding owner information
        tables = soup.find_all(
                'table', {'class', 'devkit-table'})
        for table in tables:
            caption = table.find('caption', {'class', 'devkit-table-title'})

            # This is the table we are interested in
            if caption.text == 'Registered Owner':
                print(table)
                rows = table.find_all('tr')
                for row in rows:
                    cols = row.find_all('td')
                    for col in cols:
                        if col['data-label'] == 'Name':
                            name = col.text
                        elif col['data-label'] == 'Street':
                            addr = addr + col.text
                        elif col['data-label'] == 'City':
                            city = col.text
                        elif col['data-label'] == 'State':
                            city = city + ', ' + col.text
                        elif col['data-label'] == 'Zip Code':
                            zip_code = col.text
        return Owner(name, addr, city, zip_code, 'USA'), Aircraft(tail_n)
    else:
        print("[!][{}] HTTP status code from {}"\
                .format(resp.status_code, resp.url))
    print('[!] Error while retrieving from US')


def IS(tail_n):
    name = ''
    street = ''
    city = ''
    s = requests.session()
    # s.mount("https://", TLSv1Adapter()) # Outdated TSLv1 -> Needs a specific adapter
    req = s.get(
        'https://www.icetra.is/aviation/aircraft/register?aq='+tail_n)
    if req.status_code == 200:
        soup = BeautifulSoup(req.text, 'html.parser')
        own = soup.find('li', {'class': 'owner'})
        if own is not None:
            won = own.stripped_strings
            for i, j in enumerate(won):
                if i == 1:
                    name = j
                elif i == 2:
                    street = j
                elif i == 3:
                    city = j
        if own is not None:
            return Owner(name, street, city, '', 'Iceland'), Aircraft(tail_n)


def BE(tail_n):
    rep = requests.get('https://es.mobilit.fgov.be/aircraft-registry/rest/aircrafts?aircraftStates=REGISTERED&registrationMark=' +
                       tail_n + '&page=0&pageSize=10&sort=REGISTRATIONMARK&sortDirection=ascending')

    if rep.status_code == 200:
        j = json.loads(rep.text)
        if len(j) > 0:
            if j[0].get('id') != '':
                rep = requests.get(
                    'https://es.mobilit.fgov.be/aircraft-registry/rest/aircrafts/'+str(j[0].get('id')))
                if rep.status_code == 200:
                    # print(rep.text)
                    j = json.loads(rep.text)
                    name = j.get('stakeHolderRoleList')[0].get('name')
                    street = j.get('stakeHolderRoleList')[
                        0].get('addresses').get('street')
                    city = j.get('stakeHolderRoleList')[
                        0].get('addresses').get('city')
                    return Owner(name, street, city, '', 'Belgium'), Aircraft(tail_n)
                else:
                    return Exception("Error retrieving from BE")


def SW(tail_n):
    data = {
        'selection': 'regno',
        'regno': tail_n,
        'part': '',
        'owner': '',
        'item': '',
        'X-Requested-With': 'XMLHttpRequest'
    }
    s = requests.session()
    rep = s.post(
        'https://sle-p.transportstyrelsen.se/extweb/en-gb/sokluftfartyg', data=data)
    if rep.status_code == 200:
        soup = BeautifulSoup(
            rep.content, features="html.parser")
        results_element = soup.find_all(
            'label', {'class': 'owner-headline'})[0]
        name = results_element.parent.find(
            'a', {'class': 'open-owner-page'}).text
        street = results_element.parent.text.split('\r\n')[
            1].strip()
        city = results_element.parent.text.split('\r\n')[
            2].strip()
        country = results_element.parent.text.split('\r\n')[
            3].strip()
        return Owner(name, street, city, '', country), Aircraft(tail_n)
    else:
        raise Exception('Error retrieving from SW')


def AT(tail_n):
    tail_n = tail_n.lstrip('OE-')
    rep = requests.get(
            'https://www.austrocontrol.at/'\
                    'lfa-publish-service/v2/oenfl/'\
                    'luftfahrzeuge?kennzeichen='+tail_n
            )
    print(rep.url)

    if rep.status_code == 200:
        j = json.loads(rep.content)
        owner_infos = [x for x in j if x['kennzeichen'] == tail_n]
        name, loc_info, country = owner_infos[0]['halter'].split('\r\n')
        city, street  = loc_info.split(', ')
        npa, city = city.split(' ')
        return Owner(name, street, city, npa, 'Austria'), Aircraft(tail_n)

def CZ(tail_n):
    SwissRegister = register_from_config("CZ")
    jsonobj = SwissRegister.request_infos(tail_n)
    for obj in jsonobj['rows']:
        if obj['registration_number'] == tail_n[3:]:
            r = requests.get('https://lr.caa.cz/api/avreg/{}'.format(obj['id']))
            if r.status_code == 200:
                j = json.loads(r.content)
                name = None
                if len(j['owners']) > 0:
                    name = j['owners'][0]['display_name']
                serial_n = j['serial_number']
                manufacturer = j['manufacturer']
                return Owner(name), Aircraft(tail_n, msn=serial_n, manufacturer=manufacturer)
            else:
                print("[!] Error while searching")
                raise Exception("Error while searching") 

def UK(tail_n):
    data = {
        'Registration': tail_n[2:]
    }
    print(
        data)
    s = requests.session()
    r = s.get(
        'https://siteapps.caa.co.uk/g-info/')
    r = s.post(
        'https://ginfoapi.caa.co.uk/api/aircraft/search', data=data)
    if r.status_code == 200:
        print(r.content)

    return Owner(name=""), Aircraft(tail_n)

def IE(tail_n):
    headers = {
        'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64; rv:68.0) Gecko/20100101 Firefox/68.0'}
    # r 		= requests.get('https://www.iaa.ie/commercial-aviation/aircraft-registration-2/latest-register-and-monthly-changes-1', headers=headers)
    url_doc = 'https://www.iaa.ie/docs/default-source/publications/aircraft-registration/30-november-2019.xlsx'
    r = requests.get(
        url_doc, headers=headers)
    if r.status_code == 200:
        with open('/tmp/book.xlsx', 'wb') as f:
            f.write(
                r.content)
            book = load_workbook(
                '/tmp/book.xlsx')
            for sheet in book:
                plane = [
                    row for row in sheet.values if row[0] == tail_n]
                if len(plane) > 0:
                    name = plane[
                        0][12]
                    addr = plane[
                        0][13]
                    return Owner(name, addr, '', '', 'Ireland'), Aircraft(tail_n)
                    raise Exception(
                        'Error retrieving from Ireland register. Tail number may be wrong')


def IM(tail_n):
    tail_n = tail_n.lstrip('M-')
    r = requests.get('https://ardis.iomaircraftregistry.com/register/search?prs_rm__ptt=1&prs_rm__tt=1&prs_as__v=2&prs_rm__v1='+tail_n+'&prs_rm__pv1='+tail_n)
    if r.status_code == 200:
        soup = BeautifulSoup(
            r.content, features="html.parser")
        link = soup.find(
            'td', {'id': 'prp__rid__1__cid__2'})
        if link is not None:
            href = link.contents[0]['href']
            r = requests.get(
                'https://ardis.iomaircraftregistry.com'+href)
            if r.status_code == 200:
                soup = BeautifulSoup(
                    r.content, features="html.parser")
                own = soup.find(
                    'span', {'id': 'prv__11__value'})
                name, infos = own.text.split(',')
                if len(infos) > 1:
                    street = infos
                    return Owner(name, street, '', '', ''), Aircraft(tail_n)

def CA(tail_n):
    tail_n = tail_n.lstrip('C-')
    r = requests.get('https://wwwapps.tc.gc.ca/'\
            'saf-sec-sur/2/ccarcs-riacc/RchSimpRes.aspx'\
            '?cn=%7c%7c&mn=%7c%7c&sn=%7c%7c&on=%7c%7c&m=%7c'+tail_n+'%7c&rfr=RchSimp.aspx')
    if r.status_code == 200:
        soup = BeautifulSoup(
            r.content, features='html.parser')
    div_owner   = soup.find('div', {'id':'dvOwnerName'})
    if div_owner is not None:
        name        = div_owner.find_all('div')[1].text.strip()
        div_addr    = soup.find('div', {'id':'divOwnerAddress'})
        addr        = div_addr.find_all('div')[1].text.strip()
        div_city    = soup.find('div', {'id':'divOwnerCity'})
        city        = div_city.find_all('div')[1].text.strip()
        return Owner(name, addr, city, '', 'Canada'), Aircraft(tail_n)
    else:
        print('[!] Error retrieving from CA register. Ensure tail number exists and try again')

def DE(tail_n):
    raise NotImplementedError(
        'German register not implemented yet')

def IT(tail_n):
    headers = {
        'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64; rv:68.0) Gecko/20100101 Firefox/68.0'}
    s = requests.Session()
    r = s.get(
        'http://gavs.it/')
    soup = BeautifulSoup(
        r.content, features="html.parser")
    csrf = soup.find(
        'input', {'name': 'csrf_test_name'})
    data = {
        'csrf_test_name': csrf['value'],
        'registration': tail_n[2:]
    }

    r = s.post(
        'https://gavs.it/rci/search_registration', 
        data=data, 
        headers=headers)
    if r.status_code == 200:
        record_url = r.headers['Refresh'].split(';')[1][4:]
        r = s.get(
            record_url)
        if r.status_code == 200:
            soup = BeautifulSoup(
                r.text, features="html.parser")
            tab_owners = soup.find(
                'div', {'id': 'htab2'})
            tab_owner = tab_owners.find_all(
                'dl', {'class': 'dl-horizontal'})[0]
            owner_name = tab_owner.find_all(
                    'dd')[-1].text
            return Owner(owner_name, '', '', '', ''), Aircraft(tail_n)
        raise Exception(
                "Could not get info from IT register")

def RO(tail_n):
    raise NotImplementedError(
        'Romanian register is the following pdf document http://www.caa.ro/media/docs/OPERATORI_AERIENI_ROMANI_28.11.2019_rom-eng.pdf. Documents parsing not implemented yet')

def HR(tail_n):
    raise NotImplementedError(
        'Croatian register is a pdf document. The document is available at https://www.ccaa.hr/english/popis-registriranih-zrakoplova_101/')

def AU(tail_n):
    """
    Australia Registry

    """
    r = requests.get(
        'https://www.casa.gov.au/search-centre/aircraft-register/'+tail_n[3:])
    if r.status_code == 200:
        soup = BeautifulSoup(
            r.content, features="html.parser")
        
        name = soup.find('div', 
                {'class':'field--name-field-registration-holder'}
                ).text.strip('Registration holder:\n')
        addr = soup.find('div', 
                {'class':'field--name-field-reg-hold-address-1'}
                ).text.strip('Address 1:\n')
        city = soup.find('div', 
                {'class':'field--name-field-tx-reg-hold-suburb'}
                ).text.strip('Suburb / City:\n')
        return Owner(name, addr, city, '', 'Australia'), Aircraft(tail_n)
    raise Exception(
            "Could not get info from AU register")

def SG(tail_n):
    r = requests.get(
            'https://www.caas.gov.sg/'\
                    'docs/default-source/'\
                    'docs---srg/fs/approval-listings/'\
                    'singapore-registered-aircraft-engine-nos---feb-2022.xlsx'
                    )
    with open('/tmp/SG_register.xlsx', 'wb') as f:
        f.write(r.content)
    wb = load_workbook('/tmp/SG_register.xlsx')
    ws = wb['Aircraft Register']
    for line in ws:
        if line[1].value == tail_n:
            return Owner(line[7].value, '', '', '', '')

    raise NotImplementedError('Singapore register is a pdf document.'\
            'The document is available at '\
            'https://www.caas.gov.sg/docs/default-source/pdf/'\
            'singapore-registered-aircraft-engine-nos---oct-2019.pdf')

def NZ(tail_n):
    r = requests.get('https://www.aviation.govt.nz/aircraft/'\
            'aircraft-registration/aircraft-register/ShowDetails/'+tail_n[3:])

    if r.status_code == 200:
        soup		= BeautifulSoup(r.text, features="html.parser")
        owner_info 	= soup.find('div', {'class': 'col-md-9'})
        name 		= owner_info.text.strip().split('\n')[1].strip()
        street 		= owner_info.text.strip().split('\n')[2].strip()
        return Owner(name, street, '', '', 'New Zealand'), Aircraft(tail_n)
    raise Exception("Could not get info from NZ register")

def BR(tail_n):
    r = requests.get('https://sistemas.anac.gov.br'\
            '/aeronaves/cons_rab_resposta.asp?textMarca=' + tail_n)
    if r.status_code == 200:
        soup 		= BeautifulSoup(r.text, features="html.parser")
        headings 	= soup.find_all('th', {'scope':'row'})
        for heading in headings:
            if 'Proprietário' in heading.text:
                name = heading.parent.td.text.strip()
                return Owner(name, '', '', '', 'Brazil'), Aircraft(tail_n)
            raise Exception("Could not get info from BR register. " + r.url + r.status_code )

def UA(tail_n):
    r = requests.get('http://avia.gov.ua/register_of_aircraft.xls')
    if r.status_code == 200:
        with open('/tmp/register.xls', 'wb') as f:
            f.write(r.content)
            book 		= xlrd.open_workbook('/tmp/register.xls')
            sheet_names	= book.sheet_names()
            xl_sheet 	= book.sheet_by_name(sheet_names[0])
            for i in range(0, xl_sheet.nrows):
                if xl_sheet.row(i)[2].value == tail_n:
                    name = xl_sheet.row(i)[9].value
                    return Owner(name, '', '', '', 'Ukraine')
                    raise Exception("Could not get info from UA register")

def TH(tail_n):
    # Register dates from 2019. TODO: find more recent one
    r = requests.get('https://www.caat.or.th/wp-content/uploads/'\
            '2019/03/Aircraft-List_18-Mar-2019-Un-controlled-EN.xlsx')
    if r.status_code == 200:
        with open('/tmp/book.xlsx', 'wb') as f:
            f.write(r.content)
            book = load_workbook('/tmp/book.xlsx')
            for sheet in book:
                infos = [row for row in sheet.values if row[1] == tail_n]
                if len(infos) > 0:
                    name = infos[0][2]
                    return Owner(name, '', '', '', 'Thailand')
                else:
                    raise Exception('Number not found in thai register')
            raise Exception("Could not get info from thai register")

def RS(tail_n):
    items = []
    for i in range(0, 2000, 100):
        r = requests.get('https://apps.cad.gov.rs/ords/dcvws'\
                '/regvaz/site/listAircraft?p_reg_id=0&p_type_ac=0'\
                '&p_manufacturer=0&p_man_code=0&p_sn=0&p_user=0'\
                '&order_by=p_reg_id.asc&offset='+str(i), verify=False)
        if r.status_code == 200:
            j = json.loads(r.content)
            items.extend(j['items'])

    for item in items:
        if item['registarska_oznaka'] == tail_n:
            own = item['korisnik']
            return Owner(own, '', '', '', 'Serbia')

def DK(tail_n):
    register = register_from_config("DK")
    infos = register.request_infos(tail_n)
    soup = BeautifulSoup(infos, 'html.parser')
    tr = soup.find('tr', {'class':'ulige'})
    td_link = tr.find('a')

    # No owner infos fallback by sending link
    return 'http://www-oy-reg.dk'+td_link['href']

def LV(tail_n):
    register = register_from_config("LV")
    book = register.request_infos(tail_n)
    infos_sheet = book["Sheet1"]
    for row in infos_sheet.values:
        if tail_n in row:
            return row

def BA(tail_n):
    register = register_from_config("BA")
    infos = register.request_infos(tail_n)
    for page in infos['pages']:
        for element in page['elements']:
            if element['type'] == 'table':
                for line in element['content']:
                    tail_n_from_file = line['content'][1]['content'][0]['content']
                    if tail_n == tail_n_from_file:
                        owner_line = line['content'][5]['content']
                        # Owner info
                        owner_name = ' '.join(
                            [owner_line[i]['content'] for i in range(0, len(owner_line))]
                        )
                        return Owner(owner_name, '', '', '', '') 

def HR(tail_n):
    register = register_from_config("HR")
    infos = register.request_infos(tail_n)
    for page in infos['pages']:
        for element in page['elements']:
            if element['type'] == 'table':
                for line in element['content']:
                    print(line)


def CY(tail_n):
    register = register_from_config("CY")
    infos = register.request_infos(tail_n)
    for page in infos['pages']:
        for element in page['elements']:
            if element['type'] == 'table':
                for line in element['content']:
                    if len(line['content']) >= 1:
                        if len(line['content'][1]['content']) >= 1:
                            reg = line['content'][1]['content'][0]['content']
                            if reg == tail_n:
                                own_cell = line['content'][7]['content']
                                own = ' '.join([own_cell[i]['content'] \
                                        for i in range(0,len(own_cell))])
                                return Owner(own, '', '', '', '')
def MV(tail_n):
    register = register_from_config("MV")
    infos = register.request_infos(tail_n)
    for page in infos['pages']:
        for element in page['elements']:
            if element['type'] == 'table':
                for line in element['content']:
                    if len(line['content']) > 2:
                        if len(line['content'][3]['content']) > 0 \
                                and line['content'][3]['type'] != 'spanned-table-cell':
                            if line['content'][3]['content'][0]['content'] == tail_n:
                                owner = line['content'][17]['content']
                                own = ' '.join(owner[i]['content'] for i in range(0, len(owner)))
                                return Owner(own, '', '', '', '')
def CAY(tail_n):
    register = register_from_config("CAY")
    infos = register.request_infos(tail_n)
    for page in infos['pages']:
        for element in page['elements']:
            if element['type'] == 'table':
                for line in element['content']:
                    print(line)
                    print('---')

def GG(tail_n):
    register = register_from_config("GG")
    infos = register.request_infos(tail_n)
    for page in infos['pages']:
        for e in page['elements']:
            if e['type'] == 'table':
                for content in e['content'][1:]:
                    line_to_get = None
                    for i, line in enumerate(content['content'][0]['content']):
                        if line['content'] == tail_n:
                            line_to_get = i
                    own = ' '.join(
                            content['content'][3]['content'][line_to_get+i]['content'] \
                                    for i in range(1, 8))
                    return Owner(own, '', '', '', 'Guernsey'), Aircraft(tail_n)
                    tail_column = content['content'][0]['content']
    return ""


def MT(tail_n):
    register = register_from_config("MT")
    infos = register.request_infos(tail_n)
    for page in infos['pages']:
        for e in page['elements']:
            if e['type'] == 'table':
                for content in e['content'][1:]:
                    if content['type'] == 'table-row':
                        row         = content
                        tail        = row['content'][2]['content'][0]['content']
                        owner_arr   = row['content'][7]['content']
                        if tail in tail_n:
                            owner_name = ' '.join(c['content'] for c in owner_arr)
                            owner_infos = owner_name.split(',')
                            name = owner_infos[0]
                            addr = owner_infos[1]
                            city = owner_infos[2]
                            return Owner(name, addr, city, '', country='Malta'), Aircraft(tail_n)
    return Owner('', country="Malta")

def MD(tail_n):
    register = register_from_config("MD")
    infos = register.request_infos(tail_n)
    for page in infos['pages']:
        for content in page['elements']:
            if content['type'] == 'table':
                for item_row in content['content']:
                    if item_row['type'] == 'table-row':
                      ##  print(item_row['content'])
                        retrieved_tail = item_row['content'][2]['content']
                        owner = item_row['content'][4]['content']
                        if len(retrieved_tail) > 0:
                            if retrieved_tail[0]['content'] == tail_n:
                                owner_name = owner[0]['content']
                                return Owner(owner_name, country='Moldova')
def BZ(tail_n):
    register = register_from_config("BZ")
    infos = register.request_infos(tail_n)
    for page in infos['pages']:
        print(page)
        for content in page['elements']:
            if content['type'] == 'table':
                for item in content['content']:
                    tail = item['content'][0]['content'][0]['content']
                    if tail == tail_n:
                        # Owner infos
                        own = item['content'][3]['content'][0]['content']
                        addr = ' '.join(
                                item['content'][4]['content']\
                                [i]['content']\
                                for i in range(0, len(item['content'][4]['content'])))
                        addr, city = addr.split(',')

                        # Aircraft infos
                        manufacturer = item['content'][1]['content'][0]['content']
                        print(own)
                        return Owner(own, addr, city=city), \
                                Aircraft(tail_n, manufacturer=manufacturer)


def VE(tail_n):
    register = register_from_config("VE")
    infos = register.request_infos(tail_n)
    print(infos)
